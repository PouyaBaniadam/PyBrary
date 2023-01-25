import os
import sys
import tkinter.font
import random
import pyperclip
import openpyxl
import pandas
import sqlite3
from tkinter import *
from tkinter import ttk
from quotes import fifty_quotes


def resource_path(relative_path):
    global base_path

    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath("Assets")

    return os.path.join(base_path, relative_path)


def app():
    """Back-End of buttons start here"""

    def sayings():
        global all_sayings

        try:
            conn = sqlite3.connect(resource_path("Sayings\\sayings.db"))
        except:
            os.mkdir(resource_path("Sayings"))
            conn = sqlite3.connect(resource_path("Sayings\\sayings.db"))

        c = conn.cursor()
        c.execute(
            """
        CREATE TABLE IF NOT EXISTS sayings
        (
        saying text,
        said_by text
        )
        """
        )

        c.executemany("INSERT INTO sayings VALUES (?, ?)", fifty_quotes)
        c.execute("SELECT * FROM sayings")
        all_sayings = c.fetchall()
        conn.commit()
        return all_sayings

    def theme_selector():
        global theme_folder

        try:
            with open("Settings\\Color Setting.bin", "r") as color_setting:
                color_value = color_setting.readline()

            theme_folder = color_value

            selected_color_value = 1
            for color_picker in range(10):
                if color_picker == selected_color_value:
                    theme_folder = selected_color_value
                else:
                    break
        except:
            try:
                os.mkdir("Settings")
            except:
                pass

            os.system("attrib +h Settings")
            open("Settings\\Color Setting.bin", "w")
            os.system("attrib +h Color Setting.bin")
            with open("Settings\\Color Setting.bin", "w") as color_setting:
                color_setting.write("4")

            with open("Settings\\Color Setting.bin", "r") as color_setting:
                color_value = color_setting.readline()

            theme_folder = color_value

            selected_color_value = 1
            for color_picker in range(11):
                if color_picker == selected_color_value:
                    theme_folder = selected_color_value
                else:
                    break

        if theme_folder == "":
            with open("Settings\\Color Setting.bin", "w") as color_setting:
                color_setting.write("1")
                theme_folder = 1
                color_setting.close()

    def default_font(font_size=15):
        try:
            with open("Settings\\Font Size Setting.bin", "r") as font_size_setting:
                font_size = int(font_size_setting.read())
        except:
            font_size = 15

        try:
            with open("Settings\\Font Setting.bin", "r") as font_setting:
                font = font_setting.read()
            if font_size:
                default_font = tkinter.font.Font(family=font, size=font_size)
            else:
                default_font = tkinter.font.Font(family=font, size=15)

        except:
            if font_size:
                default_font = tkinter.font.Font(family="Gabriola", size=font_size)
            else:
                default_font = tkinter.font.Font(family="Gabriola", size=15)

        return default_font

    def create_excel_for_main_xlsx_file(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "book_name"
        sheet["B1"] = "author"
        sheet["C1"] = "release_date"
        sheet["D1"] = "bar_code"
        sheet["E1"] = "how_many_left"
        sheet["F1"] = "how_many_sold"

        try:
            workbook.save(path)
        except:
            os.mkdir("DataBase")
            workbook.save(path)

    def read_from_excel():
        global all_rows
        global book
        global sheet

        try:
            with open("Settings\\Filename Setting.bin", "r") as new_file_name_setting:
                new_file_name = new_file_name_setting.read()
        except FileNotFoundError:
            with open("Settings\\Filename Setting.bin", "w") as new_file_name_setting:
                new_file_name_setting.write("LibraryDataBase")
            with open("Settings\\Filename Setting.bin", "r") as new_file_name_setting:
                new_file_name = new_file_name_setting.read()

        if new_file_name == "LibraryDataBase":
            try:
                book = openpyxl.load_workbook("DataBase\\LibraryDataBase.xlsx")
            except:
                create_excel_for_main_xlsx_file("DataBase\\LibraryDataBase.xlsx")
                book = openpyxl.load_workbook("DataBase\\LibraryDataBase.xlsx")
        else:
            try:
                book = openpyxl.load_workbook(f"DataBase\\{new_file_name}.xlsx")
            except:
                create_excel_for_main_xlsx_file(f"DataBase\\{new_file_name}.xlsx")
                book = openpyxl.load_workbook(f"DataBase\\{new_file_name}.xlsx")

        sheet = book.active
        rows = sheet.rows

        try:
            headers = [cell.value for cell in next(rows)]
        except:

            try:
                if new_file_name == "LibraryDataBase":
                    create_excel_for_main_xlsx_file("DataBase\\LibraryDataBase.xlsx")
                    headers = [cell.value for cell in next(rows)]
                else:
                    create_excel_for_main_xlsx_file(f"DataBase\\{new_file_name}.xlsx")
                    headers = [cell.value for cell in next(rows)]

            except:
                pass

        all_rows = []

        for row in rows:
            data = {}
            for book, cell in zip(headers, row):
                data[book] = cell.value
            all_rows.append(data)

            global all_bar_codes
            global all_books

            all_bar_codes = []
            all_books = []

            for bar_codes in all_rows:
                all_bar_codes.append(str(bar_codes["bar_code"]))

            for books_name in all_rows:
                all_books.append(str(books_name["book_name"]))

    def update():
        global final_all_books
        global final_all_bar_codes

        try:
            for books_name in all_rows:
                all_books.append(books_name["book_name"])
            for bar_codes in all_rows:
                all_bar_codes.append(str(bar_codes["bar_code"]))
        except:
            pass

        try:
            final_all_books = []
            [final_all_books.append(str(x)) for x in all_books if x not in final_all_books]

            final_all_bar_codes = []
            [final_all_bar_codes.append(str(x)) for x in all_bar_codes if x not in final_all_bar_codes]
        except:
            pass

    def status_announcement(text: str, x_cordinate=250, y_cordinate=450, fg_color: str = "#FF0000"):
        message_label = Label(root, text=text, border=0, background="#EBF1F7", font=default_font(),
                              fg=fg_color)
        message_label.bind('<Configure>', lambda event: message_label.config(wraplength=message_label.winfo_width()))

        message_label.place(x=x_cordinate, y=y_cordinate, width=200, anchor=CENTER)

    def piece_of_paper():
        global piece_of_paper

        piece_of_paper = PhotoImage(file=resource_path(f"Images\\piece_of_paper.png"))
        piece_of_paper_label = Label(image=piece_of_paper, background="#0F0F0F")
        piece_of_paper_label.place(x=110, y=146)

    def piece_of_paper_with_info_on_it(add_section=False):
        global piece_of_paper_with_info_on_it

        piece_of_paper_with_info_on_it = PhotoImage(file=resource_path(f"Images\\piece_of_paper.png"))
        piece_of_paper_with_info_on_it_label = Label(image=piece_of_paper_with_info_on_it, background="#0F0F0F")
        piece_of_paper_with_info_on_it_label.place(x=110, y=146)

        if add_section is False:
            book_name_text_label = Label(root, text="Book name :", border=0, background="#EBF1F7",
                                         font=default_font(), fg="#000000")
            book_name_text_label.place(x=155, y=250)

            author_text_label = Label(root, text="Author :", border=0, background="#EBF1F7",
                                      font=default_font(), fg="#000000")
            author_text_label.place(x=155, y=280)

            release_date_text_label = Label(root, text="Release date :", border=0, background="#EBF1F7",
                                            font=default_font(), fg="#000000")
            release_date_text_label.place(x=155, y=310)

            bar_code_text_label = Label(root, text="Bar code :", border=0, background="#EBF1F7",
                                        font=default_font(), fg="#000000")
            bar_code_text_label.place(x=155, y=340)

            how_many_left_text_label = Label(root, text="Left :", border=0, background="#EBF1F7",
                                             font=default_font(), fg="#000000")
            how_many_left_text_label.place(x=155, y=370)

            how_many_sold_text_label = Label(root, text="Sold :", border=0, background="#EBF1F7",
                                             font=default_font(), fg="#000000")
            how_many_sold_text_label.place(x=155, y=400)

        else:
            book_name_text_label = Label(root, text="Book name :", border=0, background="#EBF1F7",
                                         font=default_font(), fg="#000000")
            book_name_text_label.place(x=155, y=250)

            author_text_label = Label(root, text="Author :", border=0, background="#EBF1F7",
                                      font=default_font(), fg="#000000")
            author_text_label.place(x=155, y=280)

            release_date_text_label = Label(root, text="Release date :", border=0, background="#EBF1F7",
                                            font=default_font(), fg="#000000")
            release_date_text_label.place(x=155, y=310)

            bar_code_text_label = Label(root, text="Bar code :", border=0, background="#EBF1F7",
                                        font=default_font(), fg="#000000")
            bar_code_text_label.place(x=155, y=340)

            book_count_text_label = Label(root, text="Book count :", border=0, background="#EBF1F7",
                                          font=default_font(), fg="#000000")
            book_count_text_label.place(x=155, y=370)

    def search():
        global search_screen_background
        global search_icon
        global search_bar_entry
        global choose_search_method

        read_from_excel()
        update()

        def search_bar():
            global search_icon
            global search_bar_entry

            search_bar_entry = Entry(root, border=0, background="#0F0F0F", fg="#EBF1F7",
                                     font=default_font())
            search_bar_entry.place(x=137, y=112, width=230, height=30)

            search_icon = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\search_icon.png"))

            real_search_icon = Button(root, image=search_icon,
                                      bg="#0F0F0F", activebackground="#0F0F0F", border=0,
                                      command=choose_search_method)
            real_search_icon.place(x=410, y=105)

        def edit():
            global edit_screen_background
            global search_icon
            global search_bar_entry
            global choose_edit_method
            global go_on_btn, rounded_back_btn, rounded_bar_code_generator_btn

            def bar_code_generator(event):
                code = random.randint(10000, 99999)
                pyperclip.copy(code)
                spam = pyperclip.paste()
                if len(bar_code_entry.get()) != 0:
                    bar_code_entry.delete(0, END)
                    bar_code_entry.insert(0, f"{code}")
                else:
                    bar_code_entry.insert(0, f"{code}")

            def submit():
                global sure_btn, not_sure_btn, rounded_back_btn

                def final_submit():
                    global rounded_home_button

                    piece_of_paper_with_info_on_it()

                    for _ in all_rows:
                        if _["bar_code"] == edit_form_data[3] or _["bar_code"] == int(edit_form_data[3]):
                            all_rows.remove(_)
                            try:
                                final_all_bar_codes.remove(_["bar_code"])
                                final_all_bar_codes.remove(int(_["bar_code"]))
                                all_bar_codes.remove(_["bar_code"])
                                all_bar_codes.remove(int(_["bar_code"]))
                            except:
                                pass

                            all_rows.append(temp_dict)

                            with open("Settings\\Filename Setting.bin", "r") as file_name_setting:
                                file_name = file_name_setting.read()

                            df = pandas.DataFrame.from_dict(all_rows)
                            df.to_excel(f"DataBase\\{file_name}.xlsx")

                            search_bar_entry.delete(0, END)

                            piece_of_paper_with_info_on_it()

                            book_name_label = Label(root, text=str(temp_dict["book_name"]), border=0,
                                                    background="#EBF1F7",
                                                    font=default_font(), fg="#7E447F")
                            book_name_label.place(x=245, y=250)

                            author_label = Label(root, text=temp_dict["author"], border=0, background="#EBF1F7",
                                                 font=default_font(),
                                                 fg="#7E447F")
                            author_label.place(x=220, y=280)

                            release_date_label = Label(root, text=temp_dict["release_date"], border=0,
                                                       background="#EBF1F7",
                                                       font=default_font(), fg="#7E447F")
                            release_date_label.place(x=260, y=310)

                            bar_code_label = Label(root, text=temp_dict["bar_code"], border=0, background="#EBF1F7",
                                                   font=default_font(), fg="#7E447F")
                            bar_code_label.place(x=230, y=340)

                            how_many_left_label = Label(root, text=temp_dict["how_many_left"], border=0,
                                                        background="#EBF1F7",
                                                        font=default_font(), fg="#7E447F")
                            how_many_left_label.place(x=200, y=370)

                            how_many_sold_label = Label(root, text=temp_dict["how_many_sold"], border=0,
                                                        background="#EBF1F7",
                                                        font=default_font(), fg="#7E447F")
                            how_many_sold_label.place(x=200, y=400)

                            update()

                            status_announcement(text="Updated successfully!", fg_color="#0DA700")

                            search_bar()

                    rounded_home_button = PhotoImage(
                        file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_home_button.png"))
                    real_rounded_home_button = Button(root, image=rounded_home_button, bg="#EBF1F7",
                                                      activebackground="#EBF1F7",
                                                      border=0,
                                                      command=app)
                    real_rounded_home_button.place(x=260, y=480)

                new_name = book_name_entry.get().strip()
                new_author = author_entry.get().strip()
                new_release_date = release_date_entry.get().strip()
                new_bar_code = bar_code_entry.get().strip()
                new_left = how_many_left_entry.get().strip()
                new_sold = how_many_sold_entry.get().strip()

                if new_name == "":
                    new_name = edit_form_data[0]

                if new_author == "":
                    new_author = edit_form_data[1]

                if new_release_date == "":
                    new_release_date = str(edit_form_data[2])

                if new_bar_code == "":
                    new_bar_code = str(edit_form_data[3])

                if new_left == "":
                    new_left = str(edit_form_data[4])

                if new_sold == "":
                    new_sold = str(edit_form_data[5])

                zero_list = []
                zero_counter = 0
                for _ in new_bar_code:
                    if _ == "0":
                        zero_list.append(_)
                        zero_counter += 1
                    else:
                        break

                zero_list_to_string = "".join(zero_list)
                new_bar_code = zero_list_to_string + new_bar_code[zero_counter:]

                temp_dict = {"book_name": new_name, "author": new_author,
                             "release_date": new_release_date,
                             "bar_code": new_bar_code, "how_many_sold": new_sold, "how_many_left": new_left}

                def rounded_back_btn():
                    global rounded_back_btn
                    rounded_back_btn = PhotoImage(
                        file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_back_button.png"))

                    real_rounded_back_btn = Button(root, image=rounded_back_btn, bg="#EBF1F7",
                                                   activebackground="#EBF1F7",
                                                   border=0,
                                                   command=edit)
                    real_rounded_back_btn.place(x=260, y=480)

                piece_of_paper_with_info_on_it()

                if new_name.isspace():
                    book_name_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                            font=default_font(), fg="#FF0000")
                    book_name_label.place(x=245, y=250)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn()

                elif new_author.isspace():
                    author_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                         font=default_font(), fg="#FF0000")
                    author_label.place(x=220, y=280)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn()

                elif new_release_date.isspace():
                    release_date_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                               font=default_font(), fg="#FF0000")
                    release_date_label.place(x=260, y=310)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn()

                elif new_bar_code.isspace():
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn()

                elif new_left.isspace():
                    how_many_left_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                                font=default_font(), fg="#FF0000")
                    how_many_left_label.place(x=200, y=370)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn()

                elif new_sold.isspace():
                    how_many_sold_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                                font=default_font(), fg="#FF0000")
                    how_many_sold_label.place(x=200, y=400)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn()

                elif str(new_bar_code) in final_all_bar_codes and str(new_bar_code) != str(edit_form_data[3]):
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)

                    status_announcement("Already taken!")
                    rounded_back_btn()

                elif not new_release_date.isdigit():
                    release_date_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                               font=default_font(), fg="#FF0000")
                    release_date_label.place(x=260, y=310)
                    status_announcement("Digits only!")
                    rounded_back_btn()

                elif not new_bar_code.isdigit():
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)
                    status_announcement("Digits only!")
                    rounded_back_btn()

                elif not new_left.isdigit():
                    how_many_left_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                                font=default_font(), fg="#FF0000")
                    how_many_left_label.place(x=200, y=370)
                    status_announcement("Digits only!")
                    rounded_back_btn()

                elif not new_sold.isdigit():
                    how_many_sold_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                                font=default_font(), fg="#FF0000")
                    how_many_sold_label.place(x=200, y=400)
                    status_announcement("Digits only!")
                    rounded_back_btn()

                elif (len(new_release_date)) != 4:
                    release_date_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                               font=default_font(), fg="#FF0000")
                    release_date_label.place(x=260, y=310)
                    status_announcement("4 Digits exactly!")
                    rounded_back_btn()

                elif len(new_bar_code) != 5:
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)
                    status_announcement("5 Digits exactly!")
                    rounded_back_btn()

                else:
                    book_name_label = Label(root, text=edit_form_data[0], border=0,
                                            background="#EBF1F7",
                                            font=default_font(), fg="#7E447F")
                    book_name_label.place(x=245, y=250)

                    author_label = Label(root, text=edit_form_data[1], border=0, background="#EBF1F7",
                                         font=default_font(),
                                         fg="#7E447F")
                    author_label.place(x=220, y=280)

                    release_date_label = Label(root, text=edit_form_data[2], border=0,
                                               background="#EBF1F7",
                                               font=default_font(), fg="#7E447F")
                    release_date_label.place(x=260, y=310)

                    bar_code_label = Label(root, text=edit_form_data[3], border=0, background="#EBF1F7",
                                           font=default_font(), fg="#7E447F")
                    bar_code_label.place(x=230, y=340)

                    how_many_left_label = Label(root, text=edit_form_data[4], border=0,
                                                background="#EBF1F7",
                                                font=default_font(), fg="#7E447F")
                    how_many_left_label.place(x=200, y=370)

                    how_many_sold_label = Label(root, text=edit_form_data[5], border=0,
                                                background="#EBF1F7",
                                                font=default_font(), fg="#7E447F")
                    how_many_sold_label.place(x=200, y=400)

                    if book_name_entry.get() == "" and author_entry.get() == "" and release_date_entry.get() == "" \
                            and bar_code_entry.get() == "" and how_many_left_entry.get() == "" and how_many_sold_entry.get() == "":
                        piece_of_paper_with_info_on_it()
                        status_announcement("Nothing has been changed!")

                        rounded_back_btn = PhotoImage(
                            file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_back_button.png"))
                        real_rounded_back_btn = Button(root, image=rounded_back_btn, bg="#EBF1F7",
                                                       activebackground="#EBF1F7", border=0,
                                                       command=edit)
                        real_rounded_back_btn.place(x=260, y=480)

                    else:
                        status_announcement("Are you sure about the changes?")

                        not_sure_btn = PhotoImage(
                            file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_not_sure_button.png"))
                        real_not_sure_btn = Button(root, image=not_sure_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                                   border=0,
                                                   command=edit)
                        real_not_sure_btn.place(x=220, y=480)

                        sure_btn = PhotoImage(
                            file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_sure_button.png"))
                        real_sure_btn = Button(root, image=sure_btn, bg="#EBF1F7", activebackground="#EBF1F7", border=0,
                                               command=final_submit)
                        real_sure_btn.place(x=300, y=480)

                navigation_bar()

            edit_screen_background = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\edit_screen.png"))
            edit_screen_background_label = Label(root, image=edit_screen_background)
            edit_screen_background_label.place(x=-10, y=0)

            piece_of_paper_with_info_on_it()

            book_name_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                    font=default_font())
            book_name_entry.place(x=240, y=260, width=155, height=18)

            author_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                 font=default_font())
            author_entry.place(x=220, y=290, width=175, height=18)

            release_date_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                       font=default_font())
            release_date_entry.place(x=250, y=320, width=145, height=18)

            bar_code_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                   font=default_font())
            bar_code_entry.place(x=225, y=350, width=145, height=18)

            how_many_left_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                        font=default_font())
            how_many_left_entry.place(x=200, y=380, width=195, height=18)

            how_many_sold_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                        font=default_font())
            how_many_sold_entry.place(x=200, y=410, width=195, height=18)

            rounded_bar_code_generator_btn = PhotoImage(
                file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_bar_code_generator_button.png"))
            real_rounded_bar_code_generator_btn = Button(root, image=rounded_bar_code_generator_btn, bg="#EBF1F7",
                                                         activebackground="#EBF1F7", border=0)
            real_rounded_bar_code_generator_btn.bind("<Enter>", bar_code_generator)
            real_rounded_bar_code_generator_btn.place(x=375, y=342)

            go_on_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_goon_button.png"))

            real_go_on_btn = Button(root, image=go_on_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                    border=0,
                                    command=submit)
            real_go_on_btn.place(x=300, y=480)

            rounded_back_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_back_button.png"))

            real_rounded_back_btn = Button(root, image=rounded_back_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                           border=0,
                                           command=search)
            real_rounded_back_btn.place(x=220, y=480)

            search_bar()

            navigation_bar()

            def new_data_get():

                global bg100, search_icon, search_bar_entry, go_on_btn, rounded_back_btn

                bg100 = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\newdataget.png"))
                bg100_label = Label(root, image=bg100)
                bg100_label.place(x=-10, y=0)

                book_name_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                        font=default_font())
                book_name_entry.place(x=255, y=250, width=154, height=18)

                author_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                     font=default_font())
                author_entry.place(x=230, y=280, width=178, height=18)

                release_date_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                           font=default_font())
                release_date_entry.place(x=270, y=310, width=138, height=18)

                bar_code_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                       font=default_font())
                bar_code_entry.place(x=245, y=340, width=130, height=18)

                how_manys_sold_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                             font=default_font())
                how_manys_sold_entry.place(x=260, y=402, width=115, height=20)

                how_manys_left_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                             font=default_font())
                how_manys_left_entry.place(x=260, y=447, width=115, height=20)

                go_on_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_goon_button.png"))

                real_go_on_btn = Button(root, image=go_on_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                        border=0,
                                        command=submit)
                real_go_on_btn.place(x=300, y=480)

                rounded_back_btn = PhotoImage(
                    file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_back_button.png"))

                real_rounded_back_btn = Button(root, image=rounded_back_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                               border=0,
                                               command=edit)
                real_rounded_back_btn.place(x=220, y=480)

                search_bar()

                navigation_bar()

        def delete():
            global delete_screen_background
            global sure_btn, not_sure_btn

            def final_submit():
                global rounded_home_btn

                for i in range(len(all_rows)):
                    if all_rows[i]["bar_code"] == edit_form_data[3]:
                        del all_rows[i]
                        break

                with open("Settings\\Filename Setting.bin", "r") as file_name_setting:
                    file_name = file_name_setting.read()

                df = pandas.DataFrame.from_dict(all_rows)
                df.to_excel(f"DataBase\\{file_name}.xlsx")

                update()

                piece_of_paper_with_info_on_it()
                status_announcement("Deleted Successfully!", fg_color="#0DA700")

                rounded_home_btn = PhotoImage(
                    file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_home_button.png"))

                real_rounded_home_btn = Button(root, image=rounded_home_btn, bg="#EBF1F7",
                                               activebackground="#EBF1F7",
                                               border=0,
                                               command=app)
                real_rounded_home_btn.place(x=260, y=480)

            delete_screen_background = PhotoImage(
                file=resource_path(f"Images\\Theme_{theme_folder}\\delete_screen.png"))
            delete_screen_background_label = Label(root, image=delete_screen_background)
            delete_screen_background_label.place(x=-10, y=0)

            piece_of_paper_with_info_on_it()

            search_bar()
            navigation_bar()

            book_name_label = Label(root, text=edit_form_data[0], border=0,
                                    background="#EBF1F7",
                                    font=default_font(), fg="#7E447F")
            book_name_label.place(x=245, y=250)

            author_label = Label(root, text=edit_form_data[1], border=0, background="#EBF1F7",
                                 font=default_font(),
                                 fg="#7E447F")
            author_label.place(x=220, y=280)

            release_date_label = Label(root, text=edit_form_data[2], border=0,
                                       background="#EBF1F7",
                                       font=default_font(), fg="#7E447F")
            release_date_label.place(x=260, y=310)

            bar_code_label = Label(root, text=edit_form_data[3], border=0, background="#EBF1F7",
                                   font=default_font(), fg="#7E447F")
            bar_code_label.place(x=230, y=340)

            how_many_left_label = Label(root, text=edit_form_data[4], border=0,
                                        background="#EBF1F7",
                                        font=default_font(), fg="#7E447F")
            how_many_left_label.place(x=200, y=370)

            how_many_sold_label = Label(root, text=edit_form_data[5], border=0,
                                        background="#EBF1F7",
                                        font=default_font(), fg="#7E447F")
            how_many_sold_label.place(x=200, y=400)

            status_announcement("Are you sure about the deletion?")

            not_sure_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_not_sure_button.png"))
            real_not_sure_btn = Button(root, image=not_sure_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                       border=0,
                                       command=search)
            real_not_sure_btn.place(x=220, y=480)

            sure_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_sure_button.png"))
            real_sure_btn = Button(root, image=sure_btn, bg="#EBF1F7", activebackground="#EBF1F7", border=0,
                                   command=final_submit)
            real_sure_btn.place(x=300, y=480)

        def choose_search_method():
            global temp_search, search_bar_entry, temp_search, search_bar_entry

            def search_by_name():
                global edit_btn, delete_btn, rounded_back_btn
                global search_bar_entry
                global edit_form_data

                name = temp_search
                temp_list = []

                for _ in all_rows:
                    temp_list.append(str(_["book_name"]))

                if name in temp_list:
                    global search_screen_background
                    search_screen_background = PhotoImage(
                        file=resource_path(f"Images\\Theme_{theme_folder}\\search_screen.png"))
                    search_screen_background_label = Label(root, image=search_screen_background)
                    search_screen_background_label.place(x=-10, y=0)

                    for g in all_rows:
                        if name == g["book_name"]:
                            break

                    book_name1 = g["book_name"]
                    author1 = g["author"]
                    release_date1 = g["release_date"]
                    bars_code1 = g["bar_code"]
                    how_many_sold1 = g["how_many_sold"]
                    how_many_left1 = g["how_many_left"]

                    edit_form_data = [book_name1, author1, release_date1, bars_code1, how_many_left1, how_many_sold1]

                    piece_of_paper_with_info_on_it()

                    book_name_label = Label(root, text=edit_form_data[0], border=0,
                                            background="#EBF1F7", font=default_font(),
                                            fg="#7E447F")
                    book_name_label.place(x=245, y=250)

                    author_label = Label(root, text=edit_form_data[1], border=0, background="#EBF1F7",
                                         font=default_font(), fg="#7E447F")
                    author_label.place(x=220, y=280)

                    release_date_label = Label(root, text=edit_form_data[2], border=0, background="#EBF1F7",
                                               font=default_font(), fg="#7E447F")
                    release_date_label.place(x=260, y=310)

                    bar_code_label = Label(root, text=edit_form_data[3], border=0, background="#EBF1F7",
                                           font=default_font(), fg="#7E447F")
                    bar_code_label.place(x=230, y=340)

                    how_many_left_label = Label(root, text=edit_form_data[4], border=0, background="#EBF1F7",
                                                font=default_font(), fg="#7E447F")
                    how_many_left_label.place(x=200, y=370)

                    how_many_sold_label = Label(root, text=edit_form_data[5], border=0, background="#EBF1F7",
                                                font=default_font(), fg="#7E447F")
                    how_many_sold_label.place(x=200, y=400)

                    delete_btn = PhotoImage(
                        file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_delete_button.png"))

                    edit_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_edit_button.png"))

                    real_delete_btn = Button(root, image=delete_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                             border=0,
                                             command=delete)
                    real_delete_btn.place(x=220, y=480)

                    real_edit_btn = Button(root, image=edit_btn, bg="#EBF1F7", activebackground="#EBF1F7", border=0,
                                           command=edit)
                    real_edit_btn.place(x=300, y=480)

                    search_bar()

                else:
                    piece_of_paper_with_info_on_it()
                    status_announcement(text="This book doesn't exists!")

                    book_name_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                            font=default_font(), fg="#FF0000")
                    book_name_label.place(x=245, y=250)

                    rounded_back_btn = PhotoImage(
                        file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_back_button.png"))

                    real_rounded_back_btn = Button(root, image=rounded_back_btn, bg="#EBF1F7",
                                                   activebackground="#EBF1F7",
                                                   border=0,
                                                   command=search)
                    real_rounded_back_btn.place(x=260, y=480)

                    search_bar()

                navigation_bar()

            def search_by_bar_code():
                global edit_btn, delete_btn, rounded_back_btn
                global search_bar_entry
                global edit_form_data

                bar_code = temp_search
                temp_list = []

                for _ in all_rows:
                    temp_list.append(str(_["bar_code"]))

                if bar_code in temp_list:
                    global search_screen_background
                    search_screen_background = PhotoImage(
                        file=resource_path(f"Images\\Theme_{theme_folder}\\search_screen.png"))
                    search_screen_background_label = Label(root, image=search_screen_background)
                    search_screen_background_label.place(x=-10, y=0)

                    for g in all_rows:
                        if bar_code == g["bar_code"]:
                            break

                    book_name1 = g["book_name"]
                    author1 = g["author"]
                    release_date1 = g["release_date"]
                    bars_code1 = g["bar_code"]
                    how_many_sold1 = g["how_many_sold"]
                    how_many_left1 = g["how_many_left"]

                    edit_form_data = [book_name1, author1, release_date1, bars_code1, how_many_left1, how_many_sold1]

                    piece_of_paper_with_info_on_it()

                    book_name_label = Label(root, text=edit_form_data[0], border=0,
                                            background="#EBF1F7", font=default_font(),
                                            fg="#7E447F")
                    book_name_label.place(x=245, y=250)

                    author_label = Label(root, text=edit_form_data[1], border=0, background="#EBF1F7",
                                         font=default_font(), fg="#7E447F")
                    author_label.place(x=220, y=280)

                    release_date_label = Label(root, text=edit_form_data[2], border=0, background="#EBF1F7",
                                               font=default_font(), fg="#7E447F")
                    release_date_label.place(x=260, y=310)

                    bar_code_label = Label(root, text=edit_form_data[3], border=0, background="#EBF1F7",
                                           font=default_font(), fg="#7E447F")
                    bar_code_label.place(x=230, y=340)

                    how_many_left_label = Label(root, text=edit_form_data[4], border=0, background="#EBF1F7",
                                                font=default_font(), fg="#7E447F")
                    how_many_left_label.place(x=200, y=370)

                    how_many_sold_label = Label(root, text=edit_form_data[5], border=0, background="#EBF1F7",
                                                font=default_font(), fg="#7E447F")
                    how_many_sold_label.place(x=200, y=400)

                    delete_btn = PhotoImage(
                        file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_delete_button.png"))

                    edit_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_edit_button.png"))

                    real_delete_btn = Button(root, image=delete_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                             border=0,
                                             command=delete)
                    real_delete_btn.place(x=220, y=480)

                    real_edit_btn = Button(root, image=edit_btn, bg="#EBF1F7", activebackground="#EBF1F7", border=0,
                                           command=edit)
                    real_edit_btn.place(x=300, y=480)

                    search_bar()

                else:
                    piece_of_paper_with_info_on_it()
                    status_announcement(text="This Bar-code doesn't exists!")

                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)

                    rounded_back_btn = PhotoImage(
                        file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_back_button.png"))

                    real_rounded_back_btn = Button(root, image=rounded_back_btn, bg="#EBF1F7",
                                                   activebackground="#EBF1F7",
                                                   border=0,
                                                   command=search)
                    real_rounded_back_btn.place(x=260, y=480)

                    search_bar()

                navigation_bar()

            temp_search = search_bar_entry.get()

            if temp_search in final_all_books:
                search_by_name()
            elif temp_search.isdigit():
                search_by_bar_code()
            else:
                search_by_name()

        search_screen_background = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\search_screen.png"))
        search_screen_background_label = Label(root, image=search_screen_background)
        search_screen_background_label.place(x=-10, y=0)

        search_bar()

        navigation_bar()

    def add():
        global add_screen_background
        global rounded_plus_btn, rounded_bar_code_generator_btn

        def bar_code_generator(event):
            code = random.randint(10000, 99999)
            pyperclip.copy(code)
            spam = pyperclip.paste()
            if len(bar_code_entry.get()) != 0:
                bar_code_entry.delete(0, END)
                bar_code_entry.insert(0, f"{code}")
            else:
                bar_code_entry.insert(0, f"{code}")

        def submit():
            global rounded_back_btn_for_errors, sure_btn, not_sure_btn
            global temporary_dict

            def rounded_back_btn_for_errors():
                global rounded_back_btn_for_errors
                rounded_back_btn_for_errors = PhotoImage(
                    file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_back_button.png"))

                real_rounded_back_btn = Button(root, image=rounded_back_btn_for_errors, bg="#EBF1F7",
                                               activebackground="#EBF1F7", border=0, command=add)
                real_rounded_back_btn.place(x=260, y=480)

            def final_submit():
                global rounded_home_button

                piece_of_paper_with_info_on_it(add_section=True)

                new_name = book_name_entry.get().strip()
                new_author = author_entry.get().strip()
                new_release_date = release_date_entry.get().strip()
                new_bar_code = bar_code_entry.get().strip()
                new_book_count = book_count_entry.get().strip()

                temporary_dict = {"book_name": str(new_name), "author": str(new_author),
                                  "release_date": str(new_release_date), "bar_code": str(new_bar_code),
                                  "how_many_left": str(new_book_count), "how_many_sold": "0"}

                all_rows.append(temporary_dict)

                all_books.append(new_name)
                final_all_books.append(new_name)
                all_bar_codes.append(new_bar_code)
                final_all_bar_codes.append(new_bar_code)

                with open("Settings\\Filename Setting.bin", "r") as file_name_setting:
                    file_name = file_name_setting.read()

                df = pandas.DataFrame.from_dict(all_rows)
                df.to_excel(f"DataBase\\{file_name}.xlsx")

                del temporary_dict

                update()

                book_name_label = Label(root, text=new_name, border=0, background="#EBF1F7",
                                        font=default_font(), fg="#7E447F")
                book_name_label.place(x=245, y=250)

                author_label = Label(root, text=new_author, border=0, background="#EBF1F7",
                                     font=default_font(), fg="#7E447F")
                author_label.place(x=220, y=280)

                release_date_label = Label(root, text=new_release_date, border=0, background="#EBF1F7",
                                           font=default_font(), fg="#7E447F")
                release_date_label.place(x=260, y=310)

                bar_code_label = Label(root, text=new_bar_code, border=0, background="#EBF1F7",
                                       font=default_font(), fg="#7E447F")
                bar_code_label.place(x=230, y=340)

                book_count_label = Label(root, text=new_book_count, border=0, background="#EBF1F7",
                                         font=default_font(), fg="#7E447F")
                book_count_label.place(x=245, y=370)

                status_announcement("Succesfully added!", fg_color="#0DA700")

                rounded_home_button = PhotoImage(
                    file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_home_button.png"))
                real_rounded_home_button = Button(root, image=rounded_home_button, bg="#EBF1F7",
                                                  activebackground="#EBF1F7",
                                                  border=0,
                                                  command=app)
                real_rounded_home_button.place(x=260, y=480)

            piece_of_paper_with_info_on_it(add_section=True)

            new_name = book_name_entry.get().strip()
            new_author = author_entry.get().strip()
            new_release_date = release_date_entry.get().strip()
            new_bar_code = bar_code_entry.get().strip()
            new_book_count = book_count_entry.get().strip()

            zero_list = []
            zero_counter = 0
            for _ in new_bar_code:
                if _ == "0":
                    zero_list.append(_)
                    zero_counter += 1
                else:
                    break

            zero_list_to_string = "".join(zero_list)
            new_bar_code = zero_list_to_string + new_bar_code[zero_counter:]

            all_bar_codes = []
            final_all_bar_codes = []
            all_books = []
            final_all_books = []

            for _ in all_rows:
                all_bar_codes.append(_["bar_code"])
                final_all_bar_codes.append(_["bar_code"])
                all_books.append(_["book_name"])
                final_all_books.append(_["book_name"])

            if new_name == "" and new_author == "" and new_release_date == "" and new_bar_code == "" and new_book_count == "":
                piece_of_paper_with_info_on_it(add_section=True)
                status_announcement("All fields are empty!")

                rounded_back_btn_for_errors = PhotoImage(
                    file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_back_button.png"))
                real_rounded_back_btn = Button(root, image=rounded_back_btn_for_errors, bg="#EBF1F7",
                                               activebackground="#EBF1F7", border=0,
                                               command=add)
                real_rounded_back_btn.place(x=260, y=480)

            else:
                if new_name == "":
                    book_name_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                            font=default_font(), fg="#FF0000")
                    book_name_label.place(x=245, y=250)

                    status_announcement("This field is needed!")
                    rounded_back_btn_for_errors()

                elif new_author == "":
                    author_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                         font=default_font(), fg="#FF0000")
                    author_label.place(x=220, y=280)

                    status_announcement("This field is needed!")
                    rounded_back_btn_for_errors()

                elif new_release_date == "":
                    release_date_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                               font=default_font(), fg="#FF0000")
                    release_date_label.place(x=260, y=310)

                    status_announcement("This field is needed!")
                    rounded_back_btn_for_errors()

                elif new_bar_code == "":
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)

                    status_announcement("This field is needed!")
                    rounded_back_btn_for_errors()

                elif new_book_count == "":
                    book_count_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                             font=default_font(), fg="#FF0000")
                    book_count_label.place(x=245, y=370)

                    status_announcement("This field is needed!")
                    rounded_back_btn_for_errors()

                elif new_name.isspace():
                    book_name_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                            font=default_font(), fg="#FF0000")
                    book_name_label.place(x=245, y=250)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn_for_errors()

                elif new_author.isspace():
                    author_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                         font=default_font(), fg="#FF0000")
                    author_label.place(x=220, y=280)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn_for_errors()

                elif new_release_date.isspace():
                    release_date_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                               font=default_font(), fg="#FF0000")
                    release_date_label.place(x=260, y=310)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn_for_errors()

                elif new_bar_code.isspace():
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn_for_errors()

                elif new_book_count.isspace():
                    book_count_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                             font=default_font(), fg="#FF0000")
                    book_count_label.place(x=245, y=370)

                    status_announcement("Empty string not allowed!")
                    rounded_back_btn_for_errors()

                elif new_bar_code in final_all_bar_codes:
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)

                    status_announcement("Already taken!")
                    rounded_back_btn_for_errors()

                elif not new_release_date.isdigit():
                    release_date_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                               font=default_font(), fg="#FF0000")
                    release_date_label.place(x=260, y=310)
                    status_announcement("Digits only!")
                    rounded_back_btn_for_errors()

                elif not new_bar_code.isdigit():
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)
                    status_announcement("Digits only!")
                    rounded_back_btn_for_errors()

                elif not new_book_count.isdigit():
                    book_count_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                             font=default_font(), fg="#FF0000")
                    book_count_label.place(x=245, y=370)
                    status_announcement("Digits only!")
                    rounded_back_btn_for_errors()

                elif (len(new_release_date)) != 4:
                    release_date_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                               font=default_font(), fg="#FF0000")
                    release_date_label.place(x=260, y=310)
                    status_announcement("4 Digits exactly!")
                    rounded_back_btn_for_errors()

                elif len(new_bar_code) != 5:
                    bar_code_label = Label(root, text="⛔", border=0, background="#EBF1F7",
                                           font=default_font(), fg="#FF0000")
                    bar_code_label.place(x=230, y=340)
                    status_announcement("5 Digits exactly!")
                    rounded_back_btn_for_errors()

                else:
                    if new_bar_code in final_all_bar_codes:
                        piece_of_paper_with_info_on_it(add_section=True)
                        status_announcement("Already taken!")

                        rounded_back_btn_for_errors()

                    else:
                        piece_of_paper_with_info_on_it(add_section=True)

                        book_name_label = Label(root, text=new_name, border=0, background="#EBF1F7",
                                                font=default_font(), fg="#7E447F")
                        book_name_label.place(x=245, y=250)

                        author_label = Label(root, text=new_author, border=0, background="#EBF1F7",
                                             font=default_font(), fg="#7E447F")
                        author_label.place(x=220, y=280)

                        release_date_label = Label(root, text=new_release_date, border=0, background="#EBF1F7",
                                                   font=default_font(), fg="#7E447F")
                        release_date_label.place(x=260, y=310)

                        bar_code_label = Label(root, text=new_bar_code, border=0, background="#EBF1F7",
                                               font=default_font(), fg="#7E447F")
                        bar_code_label.place(x=230, y=340)

                        book_count_label = Label(root, text=new_book_count, border=0, background="#EBF1F7",
                                                 font=default_font(), fg="#7E447F")
                        book_count_label.place(x=245, y=370)

                        status_announcement(text="Are you sure you wanna add it?")

                        not_sure_btn = PhotoImage(
                            file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_not_sure_button.png"))
                        real_not_sure_btn = Button(root, image=not_sure_btn, bg="#EBF1F7", activebackground="#EBF1F7",
                                                   border=0,
                                                   command=submit)
                        real_not_sure_btn.place(x=220, y=480)

                        sure_btn = PhotoImage(
                            file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_sure_button.png"))
                        real_sure_btn = Button(root, image=sure_btn, bg="#EBF1F7", activebackground="#EBF1F7", border=0,
                                               command=final_submit)
                        real_sure_btn.place(x=300, y=480)

        add_screen_background = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\add_screen.png"))
        add_screen_background_label = Label(root, image=add_screen_background)
        add_screen_background_label.place(x=-10, y=0)

        piece_of_paper_with_info_on_it(add_section=True)

        book_name_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                font=default_font())
        book_name_entry.place(x=240, y=260, width=155, height=18)

        author_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                             font=default_font())
        author_entry.place(x=220, y=290, width=175, height=18)

        release_date_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                   font=default_font())
        release_date_entry.place(x=250, y=320, width=145, height=18)

        bar_code_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                               font=default_font())
        bar_code_entry.place(x=225, y=350, width=145, height=18)

        book_count_entry = Entry(root, border=0, background="#D9BC9B", fg="#000000",
                                 font=default_font())
        book_count_entry.place(x=245, y=380, width=150, height=18)

        rounded_bar_code_generator_btn = PhotoImage(
            file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_bar_code_generator_button.png"))
        real_rounded_bar_code_generator_btn = Button(root, image=rounded_bar_code_generator_btn, bg="#EBF1F7",
                                                     activebackground="#EBF1F7", border=0)
        real_rounded_bar_code_generator_btn.bind("<Enter>", bar_code_generator)
        real_rounded_bar_code_generator_btn.place(x=375, y=342)

        rounded_plus_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_plus_button.png"))
        real_rounded_plus_btn = Button(root, image=rounded_plus_btn, bg="#EBF1F7", activebackground="#EBF1F7", border=0,
                                       command=submit)
        real_rounded_plus_btn.place(x=260, y=480)

        navigation_bar()

    def exit():
        root.destroy()

    def setting():
        global setting_screen
        global theme_1_preview, theme_2_preview, theme_3_preview, theme_4_preview, theme_5_preview, theme_6_preview, \
            theme_7_preview, theme_8_preview, theme_9_preview, theme_10_preview
        global sure_btn, not_sure_btn, select_font_btn

        def submit(value):
            with open("Settings\\Color Setting.bin", "w") as color_setting:
                color_setting.write(str(value))

            new_file_name = new_file_name_entry.get().strip()

            if new_file_name != "":
                with open("Settings\\Filename Setting.bin", "r") as file_name_setting:
                    old_file_name = file_name_setting.read()

                os.rename(f"DataBase\\{old_file_name}.xlsx", f"DataBase\\{new_file_name}.xlsx")

                with open(f"Settings\\Filename Setting.bin", "w") as file_name_setting:
                    file_name_setting.write(new_file_name)

            app()

        def select_font():

            def font_test(font):
                selected_font.config(family=font_list_box.get(font_list_box.curselection()))

            def font_size_submit(event):
                global temp_font_size
                temp_font_size = font_size_combo.get()

            def font_submit():
                try:
                    selected_font = font_list_box.get(font_list_box.curselection())
                except:
                    try:
                        with open("Settings\\Font Setting.bin", "r") as font_setting:
                            selected_font = font_setting.read()
                    except FileNotFoundError:
                        selected_font = "Gabriola"

                if selected_font == "None":
                    try:
                        with open("Settings\\Font Setting.bin", "r") as font_setting:
                            selected_font = font_setting.read()
                    except FileNotFoundError:
                        selected_font = "None"

                    with open("Settings\\Font Setting.bin", "w") as font_setting:
                        font_setting.write(selected_font)
                else:
                    with open("Settings\\Font Setting.bin", "w") as font_setting:
                        font_setting.write(selected_font)

                try:
                    font_size = temp_font_size
                except NameError:
                    font_size = "15"

                with open("Settings\\Font Size Setting.bin", "w") as font_size_setting:
                    font_size_setting.write(font_size)

                top.destroy()
                setting()

            top = Toplevel()
            top.title("Font selector")
            top.geometry("475x570")
            top.iconbitmap(resource_path("Images\\font_icon.ico"))

            try:
                with open("Settings\\Font Setting.bin", "r") as font_setting:
                    font = font_setting.read()
                selected_font = tkinter.font.Font(family=font, size=15)

            except:
                selected_font = tkinter.font.Font(family="Gabriola", size=15)

            font_test_frame = Frame(top, width=450, height=250)
            font_test_frame.pack(pady=10)
            font_test_frame.grid_propagate(False)
            font_test_frame.columnconfigure(0, weight=1)

            text = Text(font_test_frame, font=selected_font)
            text.grid(row=0, column=0)
            text.grid_rowconfigure(0, weight=1)
            text.grid_columnconfigure(0, weight=1)

            scrollbar = Scrollbar(top)
            scrollbar.pack(side=RIGHT, fill=Y)

            font_list_box = Listbox(top, selectmode=SINGLE, width=100, height=15)
            for font in tkinter.font.families():
                font_list_box.insert("end", font)

            font_list_box.pack(side=TOP, fill=BOTH)
            scrollbar.config(command=font_list_box.yview)

            font_list_box.bind("<ButtonRelease-1>", font_test)

            available_font_sizes = ["10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20"]

            try:
                with open("Settings\\Font Size Setting.bin", "r") as font_size_setting:
                    font_size = font_size_setting.read()
            except:
                font_size = "15"

            font_size_combo = ttk.Combobox(top, values=available_font_sizes)
            font_size_combo.set(font_size)
            font_size_combo.bind("<<ComboboxSelected>>", font_size_submit)
            font_size_combo.pack()

            font_submit_btn = Button(top, text="Save font", command=font_submit)
            font_submit_btn.pack()

        setting_screen = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\setting_screen.png"))
        setting_screen_label = Label(root, image=setting_screen)
        setting_screen_label.place(x=-10, y=0)

        new_file_name_entry = Entry(root, border=0, background="#0F0F0F", fg="#EBF1F7",
                                    font=default_font())
        new_file_name_entry.place(x=145, y=163, width=238, height=30)

        variable = IntVar()
        try:
            with open("Settings\\Color Setting.bin", "r") as color_setting:
                selected_theme = int(color_setting.readline())
        except:
            selected_theme = 1

        variable.set(selected_theme)

        theme_1_preview = PhotoImage(file=resource_path(f"Images\\Theme_1\\theme_preview.png"))
        theme_1_radio_button = Radiobutton(root, image=theme_1_preview, variable=variable, value=1,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_1_radio_button.place(x=80, y=275)

        theme_2_preview = PhotoImage(file=resource_path(f"Images\\Theme_2\\theme_preview.png"))
        theme_2_radio_button = Radiobutton(root, image=theme_2_preview, variable=variable, value=2,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_2_radio_button.place(x=160, y=275)

        theme_3_preview = PhotoImage(file=resource_path(f"Images\\Theme_3\\theme_preview.png"))
        theme_3_radio_button = Radiobutton(root, image=theme_3_preview, variable=variable, value=3,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_3_radio_button.place(x=240, y=275)

        theme_4_preview = PhotoImage(file=resource_path(f"Images\\Theme_4\\theme_preview.png"))
        theme_4_radio_button = Radiobutton(root, image=theme_4_preview, variable=variable, value=4,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_4_radio_button.place(x=320, y=275)

        theme_5_preview = PhotoImage(file=resource_path(f"Images\\Theme_5\\theme_preview.png"))
        theme_5_radio_button = Radiobutton(root, image=theme_5_preview, variable=variable, value=5,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_5_radio_button.place(x=400, y=275)

        theme_6_preview = PhotoImage(file=resource_path(f"Images\\Theme_6\\theme_preview.png"))
        theme_6_radio_button = Radiobutton(root, image=theme_6_preview, variable=variable, value=6,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_6_radio_button.place(x=80, y=335)

        theme_7_preview = PhotoImage(file=resource_path(f"Images\\Theme_7\\theme_preview.png"))
        theme_7_radio_button = Radiobutton(root, image=theme_7_preview, variable=variable, value=7,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_7_radio_button.place(x=160, y=335)

        theme_8_preview = PhotoImage(file=resource_path(f"Images\\Theme_8\\theme_preview.png"))
        theme_8_radio_button = Radiobutton(root, image=theme_8_preview, variable=variable, value=8,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_8_radio_button.place(x=240, y=335)

        theme_9_preview = PhotoImage(file=resource_path(f"Images\\Theme_9\\theme_preview.png"))
        theme_9_radio_button = Radiobutton(root, image=theme_9_preview, variable=variable, value=9,
                                           background="#0F0F0F",
                                           activebackground="#0F0F0F")
        theme_9_radio_button.place(x=320, y=335)

        theme_10_preview = PhotoImage(file=resource_path(f"Images\\Theme_10\\theme_preview.png"))
        theme_10_radio_button = Radiobutton(root, image=theme_10_preview, variable=variable, value=10,
                                            background="#0F0F0F",
                                            activebackground="#0F0F0F")
        theme_10_radio_button.place(x=400, y=335)

        select_font_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\select_font_btn.png"))
        real_select_font_btn = Button(root, image=select_font_btn, bg="#0F0F0F", activebackground="#0F0F0F", border=0,
                                      command=select_font)
        real_select_font_btn.place(x=210, y=410)

        not_sure_btn = PhotoImage(
            file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_not_sure_button.png"))
        real_not_sure_btn = Button(root, image=not_sure_btn, bg="#0F0F0F", activebackground="#0F0F0F", border=0)
        real_not_sure_btn.place(x=180, y=500)

        sure_btn = PhotoImage(
            file=resource_path(f"Images\\Theme_{theme_folder}\\rounded_sure_button.png"))
        real_sure_btn = Button(root, image=sure_btn, bg="#0F0F0F", activebackground="#0F0F0F", border=0,
                               command=lambda: submit(variable.get()))
        real_sure_btn.place(x=300, y=500)

        navigation_bar()

    def navigation_bar():
        global home_btn, search_btn, add_btn, delete_btn, edit_btn, barcodegenerator_btn, latestupdate_btn, \
            bookcounter_btn, exit_btn, setting_btn

        home_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\home_button.png"))

        search_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\search_button.png"))

        add_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\add_button.png"))

        setting_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\setting_button.png"))

        real_home_btn = Button(root, image=home_btn, bg="#222222", activebackground="#222222",
                               border=0,
                               command=app)
        real_home_btn.place(x=5, y=100)

        real_search_btn = Button(root, image=search_btn, bg="#222222", activebackground="#222222",
                                 border=0,
                                 command=search)

        real_search_btn.place(x=5, y=150)

        real_add_btn = Button(root, image=add_btn, bg="#222222", activebackground="#222222", border=0,
                              command=add)
        real_add_btn.place(x=5, y=200)

        real_setting_btn = Button(root, image=setting_btn, bg="#222222",
                                  activebackground="#222222",
                                  border=0, command=setting)
        real_setting_btn.place(x=5, y=250)

        exit_btn = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\exit_button.png"))

        real_exit_btn = Button(root, image=exit_btn, bg="#222222", width=60, height=60,
                               activebackground="#222222", border=0, command=exit)
        real_exit_btn.place(x=7, y=485)

    """Back-End of ends here"""

    sayings()
    theme_selector()

    main_menu_screen = PhotoImage(file=resource_path(f"Images\\Theme_{theme_folder}\\back_ground.png"))
    main_menu_screen_label = Label(root, image=main_menu_screen)
    main_menu_screen_label.place(x=-10, y=0)

    piece_of_paper()

    saying_choice = random.randint(0, 50)

    saying_label = Label(root, text=all_sayings[saying_choice][0], border=0, background="#EBF1F7",
                         font=default_font(), fg="#000000")
    saying_label.bind('<Configure>', lambda event: saying_label.config(wraplength=saying_label.winfo_width()))
    saying_label.place(x=160, y=250, width=250)

    said_by_label = Label(root, text=all_sayings[saying_choice][1], border=0, background="#EBF1F7",
                          font=default_font(), fg="#000000")
    said_by_label.bind('<Configure>', lambda event: saying_label.config(wraplength=saying_label.winfo_width()))
    said_by_label.place(x=190, y=450, width=220)

    navigation_bar()
    read_from_excel()
    update()

    root.mainloop()


root = Tk()
root.title('PyBrary')
root.maxsize(width=475, height=570)
root.minsize(width=475, height=570)
root.iconbitmap(resource_path("Images\\PyBrary_icon.ico"))

app()

root.mainloop()
